from openpyxl import Workbook
from openpyxl.styles import Alignment


def write_to_excel(metrics: dict,
                   save_path: str) -> None:
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 写入表头（字典的键作为列标题）
    headers = list(metrics.keys())
    ws.append(headers)

    # 区分【列表型】与【单值型】key
    list_keys = []
    single_keys = []

    for k, v in metrics.items():
        if isinstance(v, list):
            list_keys.append(k)
        else:
            single_keys.append(k)

    # 假设所有列表的长度相同，写入每一行数据
    n = len(metrics[list_keys[0]])
    for i in range(n):
        row_data = []
        for header in headers:
            if header in list_keys:
                row_data.append(metrics[header][i])
            else:
                row_data.append(None)
        ws.append(row_data)

        # row_data = [metrics[header][i] for header in headers]
        # ws.append(row_data)
    
    for key in single_keys:
        # 找到它在表头中的列位置
        col_idx = headers.index(key) + 1

        # 合并范围起止行
        start_row = 2
        end_row = n + 1

        # 合并单元格
        ws.merge_cells(
            start_row=start_row,
            start_column=col_idx,
            end_row=end_row,
            end_column=col_idx
        )

        # 在合并后的首个单元格写入该单值
        ws.cell(row=start_row, column=col_idx).value = metrics[key]

    # 创建一个居中对齐的格式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 为所有单元格设置居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # 保存为 Excel 文件
    wb.save(save_path)
